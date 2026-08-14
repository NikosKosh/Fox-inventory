from openpyxl import load_workbook
from django.test import TestCase

from .models import Cabinet, Category, Employee, Equipment, Location, Organization, Room
from .services import equipment_export_workbook


class Release151ExportTests(TestCase):
    def setUp(self):
        self.owner = Organization.objects.create(name="ООО Владелец 151", prefix="R151")
        self.employee_org = Organization.objects.create(name="ООО Работодатель 151", prefix="E151")
        self.category = Category.objects.create(name="Ноутбук 151", code="N151")
        self.location = Location.objects.create(
            organization=self.owner,
            address="г. Ростов-на-Дону, ул. Тестовая, 151",
            label="Офис 151",
        )
        self.room = Room.objects.create(location=self.location, name="Кабинет 151")
        self.cabinet = Cabinet.objects.create(location=self.location, room=self.room, name="Шкаф 151")
        self.employee = Employee.objects.create(
            full_name="Иванов Иван Иванович",
            organization=self.employee_org,
        )
        self.item = Equipment.objects.create(
            internal_code="R151-N151-001",
            accounting_group=Equipment.AccountingGroup.EMPLOYEE,
            category=self.category,
            name="Ноутбук",
            manufacturer="Lenovo",
            model="ThinkPad T14",
            serial_number="SN-151",
            mac_address="AA:BB:CC:DD:EE:15",
            hostname="r151-notebook",
            owner=self.owner,
            responsible_employee=self.employee,
            location=self.location,
            room=self.room,
            cabinet=self.cabinet,
            freeform_location="Стол 15",
            quantity=1,
            usage_status=Equipment.UsageStatus.EMPLOYEE,
            condition=Equipment.Condition.USED,
            network_address="10.151.0.15",
            network_username="operator",
            notes="[SERVICE_MARKER] Источник: тестовый документ.",
        )

    def workbook(self):
        stream = equipment_export_workbook(Equipment.objects.filter(pk=self.item.pk))
        return load_workbook(stream)

    def test_export_has_readable_main_report_and_separate_detail_sheets(self):
        wb = self.workbook()
        self.assertEqual(wb.sheetnames, ["Оборудование", "Технические данные", "Служебные данные"])

        ws = wb["Оборудование"]
        self.assertEqual(ws["A1"].value, "FOX Inventory — отчёт по оборудованию")
        self.assertEqual(ws["A4"].value, "1\nПОЗИЦИЙ")
        self.assertEqual(ws["C4"].value, "1\nЕДИНИЦ")
        self.assertEqual(ws.freeze_panes, "A9")
        self.assertEqual(ws.auto_filter.ref, "A8:J9")

        headers = [ws.cell(row=8, column=column).value for column in range(1, 11)]
        self.assertEqual(headers, [
            "Код", "Контур", "Категория", "Оборудование", "Модель / серийный №",
            "Владелец", "Где находится / у кого", "Статус", "Состояние", "Кол-во",
        ])
        self.assertNotIn("Комментарий", headers)
        self.assertNotIn("MAC-адрес", headers)
        self.assertEqual(ws["E9"].value, "Lenovo ThinkPad T14\nS/N: SN-151")
        self.assertIn("Иванов Иван Иванович", ws["G9"].value)
        self.assertIn("Офис 151", ws["G9"].value)
        self.assertIn("Кабинет 151", ws["G9"].value)
        self.assertIn("Шкаф 151", ws["G9"].value)
        self.assertIn("Стол 15", ws["G9"].value)
        self.assertNotIn("SERVICE_MARKER", " ".join(str(ws.cell(row=9, column=col).value or "") for col in range(1, 11)))

        tech = wb["Технические данные"]
        self.assertEqual(tech["G5"].value, "AA:BB:CC:DD:EE:15")
        self.assertEqual(tech["H5"].value, "r151-notebook")
        self.assertEqual(tech["I5"].value, "10.151.0.15")
        self.assertEqual(tech["J5"].value, "operator")
        self.assertEqual(tech.freeze_panes, "A5")

        service = wb["Служебные данные"]
        self.assertEqual(service["K5"].value, "[SERVICE_MARKER] Источник: тестовый документ.")
        self.assertEqual(service.freeze_panes, "A5")

    def test_main_report_uses_fixed_reasonable_column_widths(self):
        wb = self.workbook()
        ws = wb["Оборудование"]
        self.assertLessEqual(ws.column_dimensions["G"].width, 40)
        self.assertLessEqual(ws.column_dimensions["D"].width, 32)
        self.assertEqual(ws.column_dimensions["J"].width, 9)
