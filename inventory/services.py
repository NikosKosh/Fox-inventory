import csv
from decimal import Decimal, InvalidOperation
from io import BytesIO, StringIO
from pathlib import Path
from django.db import transaction
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from .models import Category, Employee, Equipment, EquipmentMovement, Location, Organization, Room
from .catalog import ensure_catalog_item, format_money, parse_note_date, parse_note_price, parse_price_source

IMPORT_COLUMNS = [
    "organization", "organization_prefix", "organization_kind", "category", "category_code", "accounting_group",
    "name", "manufacturer", "model", "serial_number", "mac_address", "internal_code", "hostname",
    "status", "condition", "employee", "position", "phone", "address", "location_label", "room", "room_type",
    "freeform_location", "quantity", "unit_price", "notes", "network_address", "network_username",
]


def _clean(value):
    return str(value).strip() if value is not None else ""


def _decimal_price(value):
    text = _clean(value).replace(" ", "").replace(",", ".")
    if not text:
        return None
    try:
        return Decimal(text).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        raise ValueError(f"некорректная цена: {value}")



TECHNICAL_CATEGORY_CODES = {"SW", "R", "AP", "CAM", "NVR", "HDD", "MB", "SFP", "CAB", "UPS", "ACS", "CT", "TL"}
TECHNICAL_KEYWORDS = (
    "коммутатор", "маршрутизатор", "точка доступа", "видеокамер", "камера видеонаблюдения",
    "видеорегистратор", "монтажная коробка", "sfp", "сетевой шкаф", "антивандальный шкаф",
    "пластиковый шкаф", "источник бесперебойного питания", "ибп", "скуд", "контроллер доступа",
    "считыватель", "кабельный тестер", "инструмент для обжима",
)


def infer_accounting_group(category_code="", category_name="", name="", model=""):
    code = _clean(category_code).upper()
    text = " ".join([_clean(category_name), _clean(name), _clean(model)]).casefold().replace("ё", "е")
    if code in TECHNICAL_CATEGORY_CODES or any(word in text for word in TECHNICAL_KEYWORDS):
        return Equipment.AccountingGroup.TECHNICAL
    return Equipment.AccountingGroup.EMPLOYEE


def read_rows(uploaded_file):
    ext = Path(uploaded_file.name).suffix.lower()
    if ext == ".xlsx":
        wb = load_workbook(uploaded_file, read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        headers = [_clean(v).lower() for v in next(rows)]
        for raw in rows:
            yield dict(zip(headers, [_clean(v) for v in raw]))
    else:
        content = uploaded_file.read().decode("utf-8-sig")
        dialect = csv.Sniffer().sniff(content[:2048], delimiters=";,	,")
        reader = csv.DictReader(StringIO(content), dialect=dialect)
        for raw in reader:
            yield {str(k).strip().lower(): _clean(v) for k, v in raw.items()}


@transaction.atomic
def import_equipment(uploaded_file, user=None):
    created = 0
    updated = 0
    errors = []
    for index, row in enumerate(read_rows(uploaded_file), start=2):
        try:
            org_name = row.get("organization") or row.get("организация")
            if not org_name:
                raise ValueError("не указана organization")
            prefix = row.get("organization_prefix") or row.get("префикс") or "ORG"
            org = Organization.objects.filter(name=org_name).first()
            if not org:
                requested_prefix = prefix.upper()
                if Organization.objects.filter(prefix=requested_prefix).exists():
                    requested_prefix = f"{requested_prefix}{index}"[:12]
                org = Organization.objects.create(
                    name=org_name, prefix=requested_prefix,
                    kind=row.get("organization_kind") or Organization.Kind.COMPANY,
                )
            category_name = row.get("category") or row.get("категория") or "Другое"
            category_code = (row.get("category_code") or row.get("код категории") or "OT").upper()
            category = Category.objects.filter(name=category_name).first() or Category.objects.filter(code=category_code).first()
            if not category:
                category = Category.objects.create(name=category_name, code=category_code)
            employee = None
            employee_name = row.get("employee") or row.get("сотрудник")
            if employee_name:
                employee, _ = Employee.objects.get_or_create(
                    full_name=employee_name,
                    organization=org,
                    defaults={"position": row.get("position", ""), "phone": row.get("phone", "")},
                )
            location = None
            address = row.get("address") or row.get("адрес")
            if address:
                location, _ = Location.objects.get_or_create(
                    organization=org,
                    address=address,
                    defaults={"label": row.get("location_label", "")},
                )
            room = None
            room_name = row.get("room") or row.get("помещение")
            if room_name and location:
                room, _ = Room.objects.get_or_create(
                    location=location, name=room_name,
                    defaults={"room_type": row.get("room_type") or Room.RoomType.OTHER},
                )
            defaults = {
                "category": category,
                "accounting_group": row.get("accounting_group") or row.get("контур учёта") or infer_accounting_group(category.code, category.name, row.get("name", ""), row.get("model", "")),
                "name": row.get("name") or row.get("наименование") or category_name,
                "manufacturer": row.get("manufacturer", ""),
                "model": row.get("model", ""),
                "serial_number": row.get("serial_number", ""),
                "mac_address": row.get("mac_address") or row.get("mac") or row.get("mac-адрес") or row.get("mac адрес") or "",
                "hostname": row.get("hostname", ""),
                "owner": org,
                "responsible_employee": employee,
                "location": location,
                "room": room,
                "freeform_location": row.get("freeform_location", ""),
                "usage_status": row.get("status") or (Equipment.UsageStatus.EMPLOYEE if employee else Equipment.UsageStatus.STOCK),
                "condition": row.get("condition") or Equipment.Condition.USED,
                "quantity": int(row.get("quantity") or 1),
                "notes": row.get("notes", ""),
                "network_address": row.get("network_address", ""),
                "network_username": row.get("network_username", ""),
            }
            explicit_price = _decimal_price(
                row.get("unit_price") or row.get("price") or row.get("цена") or row.get("учётная цена")
            )
            note_text = defaults["notes"]
            imported_price = explicit_price if explicit_price is not None else parse_note_price(note_text)
            catalog_item = ensure_catalog_item(
                category=category,
                accounting_group=defaults["accounting_group"],
                name=defaults["name"],
                manufacturer=defaults["manufacturer"],
                model=defaults["model"],
                unit_price=imported_price,
                source=(parse_price_source(note_text) if note_text else f"Импорт файла {getattr(uploaded_file, 'name', '')}") or "Импорт",
                effective_date=(parse_note_date(note_text) or timezone.localdate()) if imported_price is not None else None,
                changed_by=user,
            )
            defaults["catalog_item"] = catalog_item
            internal_code = row.get("internal_code", "")
            serial = defaults["serial_number"]
            if internal_code:
                obj, was_created = Equipment.objects.update_or_create(internal_code=internal_code, defaults=defaults)
            elif serial:
                obj, was_created = Equipment.objects.update_or_create(owner=org, serial_number=serial, defaults=defaults)
            else:
                obj = Equipment.objects.create(**defaults)
                was_created = True
            EquipmentMovement.objects.create(
                equipment=obj,
                movement_type=EquipmentMovement.MovementType.IMPORT,
                to_employee=employee,
                to_organization=org,
                to_status=obj.usage_status,
                notes=f"Импорт, строка {index}",
                created_by=user,
            )
            created += int(was_created)
            updated += int(not was_created)
        except Exception as exc:
            errors.append(f"Строка {index}: {exc}")
    return created, updated, errors


REPORT_DARK = "172033"
REPORT_HEADER = "26354A"
REPORT_ACCENT = "E98A2E"
REPORT_SOFT = "F3F5F8"
REPORT_BORDER = "D9DEE7"
REPORT_TEXT = "172033"
REPORT_MUTED = "697386"
REPORT_WHITE = "FFFFFF"
REPORT_ROW_ALT = "F8FAFC"

STATUS_FILLS = {
    Equipment.UsageStatus.STOCK: "E8EEF6",
    Equipment.UsageStatus.EMPLOYEE: "E5F5EC",
    Equipment.UsageStatus.OBJECT: "E7F1FA",
    Equipment.UsageStatus.REPAIR: "FDE8D7",
    Equipment.UsageStatus.RESERVE: "FFF3CD",
    Equipment.UsageStatus.WAITING_DISPOSAL: "FCE2E2",
    Equipment.UsageStatus.DISPOSED: "E5E7EB",
    Equipment.UsageStatus.LOANED: "EEE7F8",
}

CONDITION_FILLS = {
    Equipment.Condition.NEW: "E5F5EC",
    Equipment.Condition.USED: "EEF1F5",
    Equipment.Condition.BROKEN: "FCE2E2",
}


def _report_border():
    side = Side(style="thin", color=REPORT_BORDER)
    return Border(left=side, right=side, top=side, bottom=side)


def _set_report_widths(ws, widths):
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width


def _style_report_title(ws, title, subtitle, last_column):
    last_letter = get_column_letter(last_column)
    ws.merge_cells(f"A1:{last_letter}1")
    ws["A1"] = title
    ws["A1"].font = Font(name="Calibri", size=18, bold=True, color=REPORT_WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=REPORT_DARK)
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells(f"A2:{last_letter}2")
    ws["A2"] = subtitle
    ws["A2"].font = Font(name="Calibri", size=10, color=REPORT_MUTED)
    ws["A2"].alignment = Alignment(vertical="center")
    ws.row_dimensions[2].height = 20

    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 90
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.45
    ws.page_margins.bottom = 0.45


def _style_header_row(ws, row, last_column):
    for cell in ws[row][:last_column]:
        cell.font = Font(name="Calibri", size=10, bold=True, color=REPORT_WHITE)
        cell.fill = PatternFill("solid", fgColor=REPORT_HEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _report_border()
    ws.row_dimensions[row].height = 30


def _style_data_area(ws, start_row, end_row, last_column, wrap_columns=()):
    if end_row < start_row:
        return
    wrap_columns = set(wrap_columns)
    border = _report_border()
    for row in range(start_row, end_row + 1):
        if (row - start_row) % 2:
            for col in range(1, last_column + 1):
                ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=REPORT_ROW_ALT)
        for col in range(1, last_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = Font(name="Calibri", size=10, color=REPORT_TEXT)
            cell.border = border
            cell.alignment = Alignment(
                vertical="top",
                horizontal="center" if col == last_column else "left",
                wrap_text=col in wrap_columns,
            )
        ws.row_dimensions[row].height = 30


def _equipment_model_summary(obj):
    model = " ".join(part for part in [obj.display_manufacturer, obj.display_model] if part).strip()
    if obj.serial_number:
        return f"{model}\nS/N: {obj.serial_number}" if model else f"S/N: {obj.serial_number}"
    return model


def _equipment_placement_summary(obj):
    parts = []
    if obj.responsible_employee:
        parts.append(obj.responsible_employee.full_name)
        employee_org = obj.responsible_employee.organization
        if employee_org and employee_org.pk != obj.owner_id:
            parts.append(str(employee_org))

    if obj.location:
        location_label = obj.location.label or obj.location.address
        if obj.location.organization_id != obj.owner_id:
            location_label = f"{obj.location.organization}: {location_label}"
        parts.append(location_label)
    if obj.room:
        parts.append(obj.room.name)
    if obj.cabinet:
        parts.append(obj.cabinet.name)
    if obj.freeform_location:
        parts.append(obj.freeform_location)

    unique_parts = []
    for part in parts:
        cleaned = _clean(part)
        if cleaned and cleaned not in unique_parts:
            unique_parts.append(cleaned)
    if unique_parts:
        return " · ".join(unique_parts)

    fallbacks = {
        Equipment.UsageStatus.STOCK: "Склад",
        Equipment.UsageStatus.RESERVE: "Резерв",
        Equipment.UsageStatus.REPAIR: "Ремонт",
        Equipment.UsageStatus.WAITING_DISPOSAL: "Ожидает списания",
        Equipment.UsageStatus.DISPOSED: "Списано",
        Equipment.UsageStatus.LOANED: "Передано другой организации",
    }
    return fallbacks.get(obj.usage_status, "")


def _add_kpi_card(ws, cell_range, value, label):
    ws.merge_cells(cell_range)
    cell = ws[cell_range.split(":", 1)[0]]
    cell.value = f"{value}\n{label.upper()}"
    cell.font = Font(name="Calibri", size=11, bold=True, color=REPORT_TEXT)
    cell.fill = PatternFill("solid", fgColor=REPORT_SOFT)
    cell.border = _report_border()
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def equipment_export_workbook(queryset):
    items = list(queryset.select_related(
        "category",
        "catalog_item",
        "owner",
        "responsible_employee__organization",
        "location__organization",
        "room",
        "cabinet",
    ))
    generated_date = timezone.localdate()
    position_count = len(items)
    unit_count = sum(obj.quantity for obj in items)
    status_units = {
        status: sum(obj.quantity for obj in items if obj.usage_status == status)
        for status, _label in Equipment.UsageStatus.choices
    }
    inventory_value = sum((obj.total_value or Decimal("0.00")) for obj in items)

    wb = Workbook()
    ws = wb.active
    ws.title = "Оборудование"
    ws.sheet_properties.tabColor = REPORT_ACCENT
    _style_report_title(
        ws,
        "FOX Inventory — отчёт по оборудованию",
        f"Активное оборудование · сформировано {generated_date:%d.%m.%Y}",
        12,
    )

    _add_kpi_card(ws, "A4:B5", position_count, "позиций")
    _add_kpi_card(ws, "C4:D5", unit_count, "единиц")
    _add_kpi_card(ws, "E4:F5", status_units.get(Equipment.UsageStatus.STOCK, 0), "на складе")
    _add_kpi_card(ws, "G4:H5", status_units.get(Equipment.UsageStatus.EMPLOYEE, 0), "у сотрудников")
    _add_kpi_card(ws, "I4:J5", status_units.get(Equipment.UsageStatus.OBJECT, 0), "на объектах")
    _add_kpi_card(ws, "K4:L5", format_money(inventory_value), "учётная стоимость")
    ws.row_dimensions[4].height = 25
    ws.row_dimensions[5].height = 25

    ws.merge_cells("A6:L6")
    ws["A6"] = (
        f"В ремонте: {status_units.get(Equipment.UsageStatus.REPAIR, 0)}   ·   "
        f"В резерве: {status_units.get(Equipment.UsageStatus.RESERVE, 0)}   ·   "
        f"Передано: {status_units.get(Equipment.UsageStatus.LOANED, 0)}   ·   "
        f"Ждёт списания: {status_units.get(Equipment.UsageStatus.WAITING_DISPOSAL, 0)}"
    )
    ws["A6"].font = Font(name="Calibri", size=10, color=REPORT_MUTED)
    ws["A6"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[6].height = 20

    main_headers = [
        "Код",
        "Контур",
        "Категория",
        "Оборудование",
        "Модель / серийный №",
        "Владелец",
        "Где находится / у кого",
        "Статус",
        "Состояние",
        "Кол-во",
        "Цена, ₽",
        "Стоимость, ₽",
    ]
    for column, header in enumerate(main_headers, start=1):
        ws.cell(row=8, column=column, value=header)
    _style_header_row(ws, 8, len(main_headers))

    for obj in items:
        ws.append([
            obj.internal_code,
            obj.get_accounting_group_display(),
            obj.category.name,
            obj.display_name,
            _equipment_model_summary(obj),
            str(obj.owner),
            _equipment_placement_summary(obj),
            obj.get_usage_status_display(),
            obj.get_condition_display(),
            obj.quantity,
            obj.unit_price,
            obj.total_value,
        ])

    main_last_row = ws.max_row
    _style_data_area(ws, 9, main_last_row, 12, wrap_columns={4, 5, 6, 7, 8})
    for row in range(9, main_last_row + 1):
        obj = items[row - 9]
        ws.cell(row=row, column=8).fill = PatternFill("solid", fgColor=STATUS_FILLS.get(obj.usage_status, REPORT_SOFT))
        ws.cell(row=row, column=9).fill = PatternFill("solid", fgColor=CONDITION_FILLS.get(obj.condition, REPORT_SOFT))
        ws.cell(row=row, column=8).font = Font(name="Calibri", size=10, bold=True, color=REPORT_TEXT)
        ws.cell(row=row, column=9).font = Font(name="Calibri", size=10, bold=True, color=REPORT_TEXT)
        ws.cell(row=row, column=10).number_format = "0"
        ws.cell(row=row, column=11).number_format = '#,##0.00 [$₽-ru-RU]'
        ws.cell(row=row, column=12).number_format = '#,##0.00 [$₽-ru-RU]'

    _set_report_widths(ws, [16, 16, 20, 31, 30, 28, 38, 24, 15, 9, 14, 16])
    ws.freeze_panes = "A9"
    ws.auto_filter.ref = f"A8:L{max(main_last_row, 8)}"
    ws.print_title_rows = "8:8"
    ws.print_area = f"A1:L{max(main_last_row, 8)}"

    technical = wb.create_sheet("Технические данные")
    technical.sheet_properties.tabColor = "5B7FA3"
    tech_headers = [
        "Код", "Категория", "Наименование", "Производитель", "Модель", "Серийный номер",
        "MAC-адрес", "Hostname", "Адрес управления", "Логин",
    ]
    _style_report_title(
        technical,
        "Технические данные",
        "Сетевые и идентификационные поля вынесены отдельно, чтобы не перегружать основной отчёт.",
        len(tech_headers),
    )
    for column, header in enumerate(tech_headers, start=1):
        technical.cell(row=4, column=column, value=header)
    _style_header_row(technical, 4, len(tech_headers))
    for obj in items:
        technical.append([
            obj.internal_code,
            obj.category.name,
            obj.display_name,
            obj.display_manufacturer,
            obj.display_model,
            obj.serial_number,
            obj.mac_address,
            obj.hostname,
            obj.network_address,
            obj.network_username,
        ])
    tech_last_row = technical.max_row
    _style_data_area(technical, 5, tech_last_row, len(tech_headers), wrap_columns={3, 5, 9})
    _set_report_widths(technical, [16, 20, 30, 20, 28, 23, 19, 24, 24, 22])
    technical.freeze_panes = "A5"
    technical.auto_filter.ref = f"A4:J{max(tech_last_row, 4)}"
    technical.print_title_rows = "4:4"

    service = wb.create_sheet("Служебные данные")
    service.sheet_properties.tabColor = "8B93A1"
    service_headers = [
        "Код", "Контур учёта", "Владелец", "Сотрудник", "Организация сотрудника", "Адрес",
        "Помещение", "Шкаф", "Место текстом", "Количество", "Комментарий",
    ]
    _style_report_title(
        service,
        "Служебные данные",
        "Исходные поля размещения и комментарии сохранены здесь для диагностики и сверки.",
        len(service_headers),
    )
    for column, header in enumerate(service_headers, start=1):
        service.cell(row=4, column=column, value=header)
    _style_header_row(service, 4, len(service_headers))
    for obj in items:
        service.append([
            obj.internal_code,
            obj.get_accounting_group_display(),
            str(obj.owner),
            obj.responsible_employee.full_name if obj.responsible_employee else "",
            str(obj.responsible_employee.organization) if obj.responsible_employee else "",
            str(obj.location) if obj.location else "",
            obj.room.name if obj.room else "",
            obj.cabinet.name if obj.cabinet else "",
            obj.freeform_location,
            obj.quantity,
            obj.notes,
        ])
    service_last_row = service.max_row
    _style_data_area(service, 5, service_last_row, len(service_headers), wrap_columns={3, 4, 5, 6, 7, 8, 9, 11})
    for row in range(5, service_last_row + 1):
        service.row_dimensions[row].height = 42
        service.cell(row=row, column=10).number_format = "0"
    _set_report_widths(service, [16, 18, 28, 28, 28, 36, 22, 24, 30, 10, 58])
    service.freeze_panes = "A5"
    service.auto_filter.ref = f"A4:K{max(service_last_row, 4)}"
    service.print_title_rows = "4:4"

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream

def import_template_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "Импорт"
    ws.append(IMPORT_COLUMNS)
    ws.append([
        "ООО Компания", "ORG", "company", "Ноутбук", "N", "employee", "Ноутбук", "Lenovo", "ThinkPad",
        "SN123", "AA:BB:CC:DD:EE:FF", "ORG-N-001", "org-n-001", "employee", "used", "Иванов Иван Иванович",
        "Инженер", "+7...", "г. Город, ул. Примерная, д. 1", "Главный офис", "Переговорная", "meeting", "", 1, "75000.00", "", "", "",
    ])
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream
