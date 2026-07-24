from io import BytesIO

from django.contrib.auth import get_user_model
from django.core.exceptions import ValidationError
from django.db import IntegrityError, transaction
from django.test import TestCase
from django.urls import reverse
from openpyxl import Workbook, load_workbook

from inventory.forms import EquipmentForm
from inventory.models import Category, Equipment, Organization
from inventory.services import equipment_export_workbook, import_equipment


class EquipmentMacAddressTests(TestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_user("operator", password="test-password")
        self.client.force_login(self.user)
        self.owner = Organization.objects.create(name="ООО Тест", prefix="TST")
        self.category = Category.objects.get(code="PH")

    def make_item(self, **kwargs):
        values = {
            "category": self.category,
            "name": "IP-телефон",
            "manufacturer": "Yealink",
            "model": "SIP-T31P",
            "owner": self.owner,
        }
        values.update(kwargs)
        return Equipment.objects.create(**values)

    def test_model_normalizes_mac_address(self):
        item = self.make_item(mac_address="aa-bb-cc-dd-ee-01")
        self.assertEqual(item.mac_address, "AA:BB:CC:DD:EE:01")

    def test_invalid_mac_address_is_rejected(self):
        with self.assertRaises(ValidationError):
            self.make_item(mac_address="AA:BB:CC:DD:EE")
        with self.assertRaises(ValidationError):
            self.make_item(mac_address="ZZ:AA:BB:CC:DD:EE:01")

    def test_nonempty_mac_address_is_unique(self):
        self.make_item(mac_address="AA:BB:CC:DD:EE:01")
        with self.assertRaises(IntegrityError):
            with transaction.atomic():
                self.make_item(mac_address="aabbccddee01")

    def test_equipment_form_normalizes_mac_address(self):
        form = EquipmentForm(data={
            "category": self.category.pk,
            "accounting_group": Equipment.AccountingGroup.EMPLOYEE,
            "internal_code": "",
            "name": "IP-телефон",
            "manufacturer": "Yealink",
            "model": "SIP-T31P",
            "serial_number": "TEST-PHONE-001",
            "mac_address": "aabbccddee01",
            "hostname": "",
            "owner": self.owner.pk,
            "responsible_employee": "",
            "location": "",
            "room": "",
            "cabinet": "",
            "freeform_location": "",
            "quantity": 1,
            "usage_status": Equipment.UsageStatus.STOCK,
            "condition": Equipment.Condition.NEW,
            "notes": "",
            "network_address": "",
            "network_username": "",
            "network_password": "",
            "archived": "",
        })
        self.assertTrue(form.is_valid(), form.errors)
        self.assertEqual(form.cleaned_data["mac_address"], "AA:BB:CC:DD:EE:01")

    def test_search_and_preview_include_mac_address(self):
        item = self.make_item(
            serial_number="TEST-PHONE-001",
            mac_address="AA:BB:CC:DD:EE:01",
            network_address="192.0.2.10",
        )
        response = self.client.get(reverse("equipment_list"), {"q": "DD:EE:01"})
        self.assertContains(response, item.internal_code)
        self.assertContains(response, "AA:BB:CC:DD:EE:01")

        response = self.client.get(reverse("equipment_preview", args=[item.pk]))
        self.assertContains(response, "MAC-адрес")
        self.assertContains(response, "AA:BB:CC:DD:EE:01")

    def test_import_and_export_preserve_mac_address(self):
        wb = Workbook()
        ws = wb.active
        ws.append([
            "organization", "organization_prefix", "category", "category_code", "name",
            "manufacturer", "model", "serial_number", "mac_address", "status",
        ])
        ws.append([
            "ООО Импорт", "IMP", "Телефон", "PH", "IP-телефон", "Yealink",
            "SIP-T31P", "TEST-PHONE-001", "aabbccddee01", "stock",
        ])
        source = BytesIO()
        wb.save(source)
        source.seek(0)
        source.name = "phones.xlsx"

        created, updated, errors = import_equipment(source, self.user)
        self.assertEqual((created, updated, errors), (1, 0, []))
        item = Equipment.objects.get(serial_number="TEST-PHONE-001")
        self.assertEqual(item.mac_address, "AA:BB:CC:DD:EE:01")

        exported = equipment_export_workbook(Equipment.objects.filter(pk=item.pk))
        exported_wb = load_workbook(exported, read_only=True, data_only=True)
        rows = list(exported_wb.active.iter_rows(values_only=True))
        self.assertIn("MAC-адрес", rows[0])
        self.assertIn("AA:BB:CC:DD:EE:01", rows[1])
